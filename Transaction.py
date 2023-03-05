import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet["a1"]

for row in range(2,5):
    cell = sheet.cell(row,3)
    price = cell.value * 0.9
    next_cell = sheet.cell(row,4)
    next_cell.value = price


values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e2")

wb.save("transactions2.xlsx")

